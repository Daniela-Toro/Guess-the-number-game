# IMPORTAR FUNCIONES NECESARIAS PARA EL PROGRAMA
from colorama import Fore, Style           # Para personalizar colores en la consola
from random import randint                 # Para generar números aleatorios
from datetime import datetime              # Para manejar fechas y horas
import openpyxl                            # Para trabajar con archivos Excel
from openpyxl import Workbook              # Para crear nuevos archivos Excel
import matplotlib.pyplot as plt            # Para generar gráficos
import matplotlib.image as mpimg           # Para mostrar imágenes
import pygame                              # Para manejar sonidos
import time                                # Para pausas en la ejecución

# FUNCIÓN PARA VALIDAR QUE UN NOMBRE CONTENGA SOLO LETRAS Y MÁS DE 3 CARACTERES
def names_validation(name):
    while not name.isalpha():
        name = input(Fore.LIGHTRED_EX + f"\U0001F534 ERROR: NOMBRE INVÁLIDO! SOLO PUEDE CONTENER LETRAS: ").upper()
    return name

# FUNCION PARA LIMPAR CONSOLA: GETPASS, CLS o CLEAR NO ME FUNCIONAN EN PyCharm, VS Code o Jupyter Notebooks
def clear_console():
    print("\n" * 100)  # Genera líneas vacías para simular una consola limpia

# FUNCIÓN PARA VALIDAR ENTRADAS NUMÉRICAS DENTRO DE UN RANGO
def validation(value, min, max):
    incorrect_attempts = 0
    while (not value.isdigit()) or (int(value) < min or int(value) > max):
        value = input(Fore.LIGHTRED_EX + f"\U0001F534 SOLO SE ADMITEN VALORES NUMÉRICOS ENTRE {min} Y {max}. INGRESE UN NUEVO VALOR: ")
        incorrect_attempts += 1
        # AVISAR AL USUARIO QUE SU ENTRADA ES INCORRECTA
        if incorrect_attempts >= 3:  # Mensaje de advertencia tras 3 intentos incorrectos
            incorrect_attempts = 0
            print(Fore.LIGHTRED_EX + f'SELECCIONÓ 3 VECES UN NÚMERO NO VÁLIDO, POR FAVOR, LEA DE NUEVO \U0001F600')
    return int(value)

# FUNCIÓN PARA MOSTRAR EL MENÚ PRINCIPAL
def Menu():
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(Fore.LIGHTBLUE_EX + "\U0001F527 MENÚ".center(70))
    print(Fore.LIGHTBLUE_EX + 'SELECCIÓN DE MODALIDAD'.center(70))
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(
        Fore.LIGHTBLUE_EX + "\t1. \U0001F3AE PARTIDA MODO SOLITARIO\n" +
        "\t2. \U0001F465 PARTIDA 2 JUGADORES\n" +
        "\t3. \U0001F4CA ESTADÍSTICAS\n" +
        "\t4. \U0001F6AA SALIR" +
        Style.RESET_ALL
    )

# FUNCIÓN PARA MOSTRAR EL SUBMENÚ Y SELECCIONAR LA DIFICULTAD
def submenu(name):
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(Fore.LIGHTBLUE_EX + "\U0001F527 SUBMENÚ".center(70))
    print(Fore.LIGHTBLUE_EX + ' SELECCIÓN DEL NIVEL DE DIFICULTAD'.center(70))
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(
        Fore.LIGHTBLUE_EX +
        "\t1. \U0001F95A FÁCIL (20 INTENTOS DISPONIBLES)" + "\n" +
        "\t2. \U0001F425 MEDIO (12 INTENTOS DISPONIBLES)" + "\n" +
        "\t3. \U0001F414 DIFÍCIL (5 INTENTOS DISPONIBLES)" +
        Style.RESET_ALL
    )
    difficulty = input(Fore.LIGHTCYAN_EX + f"\u27A1 {name} SELECCIONA UNA DE LAS OPCIONES ANTERIORES: ")
    difficulty = validation(difficulty, 1, 3)
    return {1: 20, 2: 12, 3: 5}[difficulty] # Retorna los intentos según la dificultad seleccionada

# # FUNCIÓN PARA REPRODUCIR SONIDO Y MOSTRAR IMAGEN DE GANAR O PERDER
def animation_game(url_photo, url_sound):
    # REPRODUCIR SONIDO
    pygame.init()
    pygame.mixer.init()
    sound = pygame.mixer.Sound(url_sound)
    sound.play()
    time.sleep(2) # Pausa para permitir la reproducción del sonido
    pygame.quit()
    # MOSTRAR IMAGEN
    img = mpimg.imread(url_photo)
    plt.imshow(img)
    plt.axis('off')
    plt.show()

# FUNCIÓN PRINCIPAL DE LÓGICA DEL JUEGO
def play_game(unknown_number, name):
    max_attempts = submenu(name)
    print(Fore.YELLOW + f"{name}, TIENES {max_attempts} INTENTOS PARA ADIVINAR!.")
    attempts = 0
    guess_number = 0
    win = False
    while (max_attempts > attempts) and (guess_number != unknown_number):
        guess_number = input(Fore.LIGHTCYAN_EX + "\u27A1 ADIVINA! INTRODUCE UN NÚMERO ENTRE 1 Y 1000: ")
        guess_number = validation(guess_number, 1, 1000)
        if guess_number == unknown_number:
            print(Fore.LIGHTYELLOW_EX + Style.BRIGHT + f"\U0001F3C6 FELICIDADES {name}! HAS GANADO!")
            animation_game('./tmp/win.jpg', './tmp/win.wav')
            win = True
        elif guess_number > unknown_number:
            print(Fore.LIGHTGREEN_EX + "EL NÚMERO A ADIVINAR ES MENOR \U00002B07")
        else:
            print(Fore.LIGHTGREEN_EX + "EL NÚMERO A ADIVINAR ES MAYOR \U00002B06")
        attempts += 1
        if win == False:
            print(Fore.YELLOW + f'\u26A0 TE QUEDAN {max_attempts - attempts} INTENTOS DISPONIBLES')
    if (max_attempts == attempts) and (win == False):
        print(Fore.RED + Style.BRIGHT + f"\U0001F534 {name}, HAS SUPERADO EL NÚMERO MÁXIMO DE INTENTOS PERMITIDOS\nSUERTE PARA LA PRÓXIMA")
        print(Fore.RED + Style.BRIGHT + '\U0001F480 GAME OVER \u2620')
        animation_game('./tmp/lose.jpg', './tmp/game_over.wav')
    save_statistics(name, attempts, win, unknown_number, max_attempts)  # Guardar estadísticas de la partida

# FUNCION PARA MODO SOLITARIO
def one_player():
    print(Fore.LIGHTBLUE_EX + "▓" * 70 + "\n" + "\U0001F3AE PARTIDA MODO SOLITARIO".center(70))
    print(Fore.LIGHTBLUE_EX + f"DESCUBRE EL NÚMERO QUE SE ESCONDE ENTRE 1 Y 1000.".center(70) + "\n" + "▓" * 70)
    unknown_number = randint(1, 1000)  # RANDINT INCLUYE TANTO 1 COMO 1000
    # print(unknown_number)
    name = names_validation(input(Fore.LIGHTCYAN_EX + "\u27A1 INTRODUCE TU NOMBRE: ").upper())
    play_game(unknown_number, name)

# FUNCION PARA MODO 2 JUGADORES
# LAS PARTIDAS GANADAS/PERDIDAS SE LE APUNTAN AL JUGADOR 2 SOLAMENTE (JUGADOR QUE ADIVINA)
def two_players():
    print(Fore.LIGHTBLUE_EX + "▓" * 70 + "\n" + "\U0001F465 PARTIDA 2 JUGADORES".center(70) + "\n" + "▓" * 70)
    print(Fore.YELLOW + "JUGADOR 1 ESCOGERÁ UN NÚMERO ENTRE 1 Y 1000\nJUGADOR 2 ADIVINARÁ EL NÚMERO Y LE CONTARÁ EL RESULTADO DE LA PARTIDA.")
    name2 = names_validation(input(Fore.LIGHTCYAN_EX + "\u27A1 JUGADOR 1, INTRODUCE TU NOMBRE: ").upper())
    name = names_validation(input(Fore.LIGHTCYAN_EX + "\u27A1 JUGADOR 2, INTRODUCE TU NOMBRE: ").upper())
    unknown_number = validation(input(Fore.LIGHTCYAN_EX + f"\u27A1 {name2}: INSERTA UN NÚMERO ENTRE 1 Y 1000: "), 1, 1000)
    clear_console()
    play_game(unknown_number, name)

# FUNCION PARA GUARDAR ESTADISTICAS
def save_statistics(name, attempts, win, unknown_number, max_attempts):
    try:
        # INTENTA ABRIR EL ARCHIVO DE ESTADÍSTICAS; SI NO EXISTE, LO CREA
        wb = openpyxl.load_workbook("GAME_STATISTICS.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "statistics"
        # AGREGA ENCABEZADOS AL ARCHIVO NUEVO
        sheet.append(["NOMBRE", "GANADOR", "NÚMERO_SECRETO", "INTENTOS_UTILIZADOS", "INTENTOS_TOTALES", "FECHA"])
        print(Fore.YELLOW + "\U0001F4C2 ARCHIVO GAME_STATISTICS.xlsx CREADO")
    try:
        # AGREGA UNA FILA CON LOS DATOS DE LA PARTIDA CUANDO EL ARCHIVO EXISTE
        dt = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        sheet.append([name, win, unknown_number, attempts, max_attempts, dt])
        wb.save("GAME_STATISTICS.xlsx")
        print(Fore.YELLOW + "\U0001F4BE RESULTADO DE LA PARTIDA GUARDADO EN ESTADÍSTICAS".center(70))
    except:
        # MENSAJE DE ERROR SI NO SE PUEDE GUARDAR
        print(Fore.RED + "\U0001F4BE \u274C NO SE HAN PODIDO GUARDAR LOS DATOS".center(70))
        print(Fore.RED + "COMPRUEBE QUE EL ARCHIVO EXCEL ESTÉ CERRADO PARA ESCRITURAR".center(70))

# FUNCION PARA MOSTRAR ESTADISTICAS
def show_statistics():
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(Fore.LIGHTBLUE_EX + "\U0001F4CA ESTADÍSTICAS".center(70))
    try:
        # INTENTA CARGAR EL ARCHIVO DE ESTADÍSTICAS
        wb = openpyxl.load_workbook("GAME_STATISTICS.xlsx")
        Hoja = wb['statistics']
        print(Fore.LIGHTBLUE_EX + "MENÚ".center(70))
        print(Fore.RED + "\u26A0 LAS PARTIDAS APARECEN GUARDADAS EN GAME_STATISTICS.xlsx")
        print(Fore.LIGHTBLUE_EX + "\t1. ESTADÍSTICAS GENERALES\n" + "\t2. ESTADÍSTICAS POR USUARIO\n" + Style.RESET_ALL)
        option = input(Fore.LIGHTCYAN_EX + "\u27A1 SELECCIONE UNA OPCIÓN: ")
        option = validation(option, 1, 2)
        user = None
        statistical_logic(Hoja, option, user)
    except FileNotFoundError:
        # MENSAJE SI NO SE ENCUENTRA EL ARCHIVO
        print(Fore.LIGHTRED_EX + "\u274C  ERROR: ARCHIVO NO ENCONTRADO\n" + "\tASEGÚRATE DE JUGAR AL MENOS UNA PARTIDA ANTES" + Style.RESET_ALL)

# FUNCIÓN CON LA LOGICA PRINCIPAL PARA RECORRER Y PROCESAR EL EXCEL
def statistical_logic(Hoja, option, user):
    if option == 2:
        # SI SE SELECCIONA ESTADÍSTICAS POR USUARIO, SOLICITA Y VALIDA EL NOMBRE
        user = names_validation(input(Fore.LIGHTCYAN_EX + '\u27A1 INTRODUZCA EL NOMBRE DEL USUARIO QUE DESEA BUSCAR: ').upper())
    players = {}
    show_header = False
    # RECORRE LAS FILAS DE LA HOJA PARA PROCESAR LOS DATOS
    for row in Hoja.iter_rows(min_row=2, values_only=True):
        name, win, unknown_number, attempts, max_attempts, date = row
        # MUESTRA ESTADÍSTICAS GENERALES O POR USUARIO SEGÚN LA OPCIÓN
        if option == 1 or (option == 2 and name == user):
            if show_header == False:
                print(Fore.GREEN + "\nESTADÍSTICAS GENERALES".center(70) + Style.RESET_ALL)
                # MUESTRA ENCABEZADOS DE LA HOJA
                for cell in Hoja[1]:
                    print(cell.value, end=" ")
                print()
                show_header = True
            # INICIALIZA LAS ESTADÍSTICAS DEL JUGADOR SI NO EXISTEN
            if name not in players:
                players[name] = {
                    "wins": 0,
                    "losses": 0
                }
            # ACTUALIZA GANANCIAS Y PÉRDIDAS
            if win == True:
                players[name]["wins"] += 1
            else:
                players[name]["losses"] += 1
            for cell in row:
                print(cell, end=" ")
            print()
    # GENERA GRÁFICAS SEGÚN LA OPCIÓN SELECCIONADA
    if option == 1:
        plot_general_statistics(players)
    else:
        plot_user_statistics(players, user)

# GRÁFICA DE ESTADÍSTICAS GENERALES
def plot_general_statistics (players):
    # CREA UNA GRÁFICA DE BARRAS PARA LAS GANANCIAS DE TODOS LOS JUGADORES
    names = list(players.keys())
    wins = []
    for name in names:
        wins.append(players[name]["wins"])
    plt.figure(figsize=(10, 6))
    plt.bar(names, wins, color='green')
    plt.xlabel('JUGADORES')
    plt.ylabel('PARTIDAS GANADAS')
    plt.title('PARTIDAS GANADAS POR JUGADOR')
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.show()

# GRÁFICA DE ESTADÍSTICAS INDIVIDUALES
def plot_user_statistics (players, user):
    # CREA UNA GRÁFICA DE BARRAS PARA LAS GANANCIAS Y PÉRDIDAS DE UN JUGADOR ESPECÍFICO
    if user in players:
        categories = ['PARTIDAS GANADAS', 'PARTIDAS PERDIDAS']
        values = [players[user]["wins"], players[user]["losses"]]
        plt.figure(figsize=(8, 6))
        plt.bar(categories, values, color=['green', 'red'])
        plt.xlabel('CATEGORÍAS')
        plt.ylabel('CANTIDAD DE PARTIDAS')
        plt.title(f'ESTADÍSTICAS DE {user}')
        plt.show()
    else:
        print(Fore.LIGHTRED_EX + f'EL USUARIO {user} NO EXISTE')

# PROGRAMA PRINCIPAL
try:
    # MOSTRAR MENSAJE DE BIENVENIDA
    print(Fore.LIGHTBLUE_EX + "▓" * 70)
    print(Fore.LIGHTBLUE_EX + Style.BRIGHT + "¡QUE COMIENCE EL JUEGO: ADIVINA EL NÚMERO! \U0001F40D \U0001F9D1\U0000200D\U0001F4BB".center(70))
    selection = "0"
    # BUCLE PRINCIPAL PARA MOSTRAR EL MENÚ HASTA QUE EL USUARIO DECIDA SALIR
    while selection != "4":
        Menu()
        selection = input(Fore.LIGHTCYAN_EX + "\u27A1 SELECCIONA UNA DE LAS OPCIONES ANTERIORES: ")
        # OPCIÓN 1: MODO SOLITARIO
        if selection == "1":
            one_player()
        # OPCIÓN 2: MODO DOS JUGADORES
        elif selection == "2":
            two_players()
        # OPCIÓN 3: MOSTRAR ESTADÍSTICAS
        elif selection == "3":
            show_statistics()
        # OPCIÓN 4: SALIR DEL JUEGO
        elif selection == "4":
            print(Fore.LIGHTRED_EX + "GRACIAS POR JUGAR, HASTA LA PRÓXIMA! \U0001F44B".center(70))
        # GESTIÓN DE ERRORES: VALOR FUERA DE LAS OPCIONES DISPONIBLES
        else:
            print(Fore.LIGHTRED_EX + "▓" * 70)
            print(Fore.LIGHTRED_EX + "\u274CERROR!:VALOR INSERTADO NO VÁLIDO\u274C".center(70))
            print(Fore.LIGHTRED_EX + "  DEBE SELECCIONAR UNA OPCIÓN ENTRE 1 Y 4".center(70))
            print(Fore.LIGHTRED_EX + "▓" * 70)
# GESTIÓN DE ERRORES GENERALES: FALLAS INESPERADAS EN LA EJECUCIÓN DEL PROGRAMA
except:
    print('\n')
    print(Fore.LIGHTRED_EX + "▓" * 70 + "\n" + "LO SENTIMOS, HA OCURRIDO ALGO INESPERADO".center(70))
    print(Fore.LIGHTRED_EX + "REINICIE EL JUEGO".center(70))
    print(Fore.LIGHTRED_EX + "▓" * 70)